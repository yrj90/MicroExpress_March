'''
Created on Mar 13, 2017

@author: hshi
'''
import caffe
import os
import numpy as np
import scipy.io as sio
from Samplers.LeaveOneSubjectOutSampler import LearveOneSubjectOutSampler
import cPickle as pickle
from DefineNet import defineNet
from DefineSolver import defineSolver
from Train import train
from Results.Tuples import Tuples
from Test import testOneNet
from sklearn import preprocessing
import shutil
import openpyxl as pyxl

def myMkdir(dirPath):
    if os.path.isdir(dirPath):
        shutil.rmtree(dirPath)
    
    os.mkdir(dirPath)

def report(sheet, currentResult):
            
    sheet['D1'] = "Best Valid Accuracy"
    sheet['D2'] = currentResult.getBestValidAccuracy()
            
    sheet['E1'] = "bestWeightInd"
    sheet['E2'] = currentResult.getBestWeightInd()
            
    sheet['F1'] = "Best Valid Weight"
    sheet['F2'] = currentResult.getBestWeightPath()
            
            
    sheet['G1'] = "Train Ratio"
    sheet['G2'] = currentResult.getRatio_train()
            
    sheet['H1'] = "Valid Ratio"
    sheet['H2'] = currentResult.getRatio_valid()
    
    sheet['I1'] = "folds num"
    sheet['I2'] = currentResult.getFoldsNum()
            
    sheet['J1'] = "Testing Accuracy of Best Valid Weight"
    sheet['J2'] = currentResult.getBestWeightTestAccuracy()
            
    sheet['K1'] = "Batch Size"
    sheet['K2'] = currentResult.getBatchSize()
            
    sheet['L1'] = "Training Epoches"
    sheet['L2'] = currentResult.getTrainingEpoches()
            
    sheet['M1'] = "Normalization Type"
    sheet['M2'] = currentResult.getNormalizationType()
            
    sheet['N1'] = "sample prenormalization"
    sheet['N2'] = currentResult.getSampleWisePreNormalization()
    

    sheet['O1'] = "minvalid Loss"

    validAccuracy = currentResult.getValidAccuracy()
    weightFilePath = currentResult.getWeightPath()
    testAccuracy = currentResult.getTestAccuracy() 
    validLoss = currentResult.getValidLoss()
            
    sheet['A1'] = "Validating Accuracy"
    sheet['B1'] = "Weight File Path"
    sheet['C1'] = "Testing Accuracy"
            
    for row in range(len(weightFilePath)):
        _ = sheet.cell(column = 1, row = row + 2, value = validAccuracy[row])
        #sheet['B'+str(row+2)] = weightFilePath[row]
        _ = sheet.cell(column = 2, row = row + 2, value = weightFilePath[row])
        #_ = sheet.cell(column = 2, row = row + 2, value = "%s" % weightFilePath[row])
        _ = sheet.cell(column = 3, row = row + 2, value = testAccuracy[row])
        
        _ = sheet.cell(column = 15, row = row + 2, value = validLoss[row])



    return sheet

def summarizeReport(sheet, result):
    sheet['A1'] = "fold_no"
    sheet['B1'] = "Best Valid Accuracy"
    sheet['C1'] = "Testing Accuracy of Best Valid Weight"
    sheet['D1'] = "bestWeightInd"        
    sheet['E1'] = "Best Valid Weight"
    
    sheet['F1'] = "Train Ratio"
    sheet['G1'] = "Valid Ratio"
    sheet['H1'] = "Batch Size"            
    sheet['I1'] = "Training Epoches"     
    sheet['J1'] = "Normalization Type"
    
    sheet['K1'] = "sample prenormalize"
    sheet['L1'] = "stateNumPERclass"
    sheet['M1'] = "bestValidLoss"
    
  

    tmpSampleNames_test = list()
    tmpPaths = list()
    
    for row in range(len(result)):
        currentResult = result[row]

                
        _ = sheet.cell(column = 1, row = row + 2, value = row + 1)
        _ = sheet.cell(column = 2, row = row + 2, value = currentResult.getBestValidAccuracy())
        _ = sheet.cell(column = 3, row = row + 2, value = currentResult.getBestWeightTestAccuracy())
        _ = sheet.cell(column = 4, row = row + 2, value = currentResult.getBestWeightInd())
        _ = sheet.cell(column = 5, row = row + 2, value = currentResult.getBestWeightPath()) 
           
        _ = sheet.cell(column = 6, row = row + 2, value = currentResult.getRatio_train())   
        _ = sheet.cell(column = 7, row = row + 2, value = currentResult.getRatio_valid())   
  
        
        _ = sheet.cell(column = 8, row = row + 2, value = currentResult.getBatchSize())          
        _ = sheet.cell(column = 9, row = row + 2, value = currentResult.getTrainingEpoches())     
        _ = sheet.cell(column = 10, row = row + 2, value = currentResult.getNormalizationType())     
        _ = sheet.cell(column = 11, row = row + 2, value = currentResult.getSampleWisePreNormalization())    
         
        _ = sheet.cell(column = 12, row = row + 2, value = currentResult.getStateTypeNumPerClass()) 
        
        
        sampleNames_test = currentResult.getSampleNames_test()
        paths = currentResult.getHMMpaths()
        
        if row == 0:
            tmpLabels = currentResult.getLabel_test()
            
        else:
            tmpLabels = np.concatenate((tmpLabels, currentResult.getLabel_test()), axis = 0)
            
        for i in range(len(sampleNames_test)):
            tmpSampleNames_test.append(sampleNames_test[i])        
            tmpPaths.append(paths[i])
            
            
        
    resultNum = len(result)
    
    for i in range(len(tmpSampleNames_test)):     
            
        _ = sheet.cell(column = 1, row = resultNum + 5 + i, value = tmpSampleNames_test[i])     
        _ = sheet.cell(column = 2, row = resultNum + 5 + i, value = tmpLabels[i])   
        
        currentPath = tmpPaths[i]
        
        for j in range(currentPath.shape[0]):  
            _ = sheet.cell(column = 3 + j, row = resultNum + 5 + i, value = currentPath[j])   
    



    return sheet



def main():
    
    caffe.set_device(2)
    caffe.set_mode_gpu()
    
    
    normalizationType = 1
    transitionEstimationType = 1
    batchSize = 100
    ratio_valid = 0.2
    trainingEpoches = 100
    DataDir = 'Data/cropped_videos_gray'
    VideoLabelDir = 'Data/Sequence_Labels/Sequence_Labels/OPR'
    FrameLabelDir = 'Data/Frame_Labels/Frame_Labels/PSPI'
    stateTypeNum_all = 17
    stateTypeNumAll = 17
    
    
    
    
    todayDate = 20170314
    rootDir = "Experiments"
    
    workingDirName = str(todayDate) + "_" + \
                     "transition(" + str(transitionEstimationType) + ")_" + \
                     "batchSize(" + str(batchSize) + ")_" + \
                     "trainingEpoches(" + str(trainingEpoches) + ")_"
                     
                     
    workingDir = os.path.join(rootDir, workingDirName)
    myMkdir(workingDir)
    
    data_all = list()
    userId_all = list()
    sequenceLabel_all = list()
    frameLabel_all = list()
    
    dirs = os.listdir(DataDir)
    
    for _, usr in enumerate(dirs):
        
        currentSampleDir = os.path.join(DataDir, usr)
      
        files = os.listdir(currentSampleDir)
        
        for _, v in enumerate(files):
            
            currentSampleName = v[0:-4]
            
            currentSampleFile = os.path.join(currentSampleDir, (currentSampleName + '.mat'))
            currentSequenceLabelFile = os.path.join(VideoLabelDir, usr, (currentSampleName + '.txt'))
            
            currentData = sio.loadmat(currentSampleFile)['faceVid']            
            
            data_all.append(currentData)
            userId_all.append(usr[0:3])
            
            with open(currentSequenceLabelFile) as f:
                currentSequenceLabel = f.read()
            sequenceLabel_all.append(int(float(currentSequenceLabel[0:-1])))
    
    
            currentFrameLabelDir = os.path.join(FrameLabelDir, usr, currentSampleName)
            
            frameLabelFiles = os.listdir(currentFrameLabelDir)
            frameNum = len(frameLabelFiles)
            
            currentFrameLabel = np.zeros([frameNum, 1])
            for i in range(frameNum):
                currentFrameLabelFile = os.path.join(currentFrameLabelDir, currentSampleName + '{:03d}'.format(i+1) + '_facs.txt')    
            
                with open(currentFrameLabelFile) as f:
                    currentSequenceLabel = f.read()
                
                currentFrameLabel[i,0] = int(float(currentSequenceLabel[0:-1]))
                    
                
            frameLabel_all.append(currentFrameLabel)

    userId = np.zeros([len(userId_all), 1])
    
    for i in range(len(userId_all)):
        userId[i] = int(userId_all[i])
                

    sampleSubject = np.concatenate((np.linspace(0, len(userId_all) - 1, len(userId_all)).reshape([len(userId_all), 1]),
                                    userId), axis = 1)
    
    
    sample = data_all
    videoLabel = sequenceLabel_all
    frameLabel = frameLabel_all
    
    
    
    
    mSampler = LearveOneSubjectOutSampler(sampleSubject, 
                                           ratio_valid,
                                           SUBJECT_INDEPENDENT_VALIDATION=False)
    
    # save data to file
    
    currentSamplingResult = mSampler.getNextSamplingResult()
    workingDirInd = 1

    result = list()
    
    while currentSamplingResult != None:
        
        currentWorkingDir = os.path.join(workingDir, str(workingDirInd))
        myMkdir(currentWorkingDir)
        
        currentResult = Tuples()
        
        

        
        sampleInd_train = currentSamplingResult.getSampleInd_train()
        sampleInd_valid = currentSamplingResult.getSampleInd_valid()
        sampleInd_test = currentSamplingResult.getSampleInd_test()
        
        
        
        sample_train = list2Array(getSample(sample, sampleInd_train), 2)
        sample_train = sample_train.reshape(-1, sample_train.shape[2]).transpose()
        
        videoLabel_train = intList2Array(getSample(videoLabel, sampleInd_train))
        frameLabel_train = getSample(frameLabel, sampleInd_train)
        clipMarker_train = list2Array(createClipMarkers(frameLabel_train), 0)
        frameLabel_train = list2Array(frameLabel_train, 0)
        
        
    
        sample_valid = list2Array(getSample(sample, sampleInd_valid),2 )
        sample_valid = sample_valid.reshape(-1, sample_valid.shape[2]).transpose()
        
        videoLabel_valid = intList2Array(getSample(videoLabel, sampleInd_valid))
        
        frameLabel_valid = getSample(frameLabel, sampleInd_valid)
        clipMarker_valid = list2Array(createClipMarkers(frameLabel_valid), 0)
        frameLabel_valid = list2Array(frameLabel_valid, 0)
        
        
        sample_test = getSample(sample, sampleInd_test)
        videoLabel_test = getSample(videoLabel, sampleInd_test)
        frameLabel_test = getSample(frameLabel, sampleInd_test)
        clipMarker_test = createClipMarkers(frameLabel_test)
        
        
        data_test = {'sample': sample_test,
                      'videoLabel': videoLabel_test,
                      'frameLabel': frameLabel_test,
                      'clipMarker': clipMarker_test}
        

        sample_test = list2Array(sample_test, 2)
        sample_test = sample_test.reshape(-1, sample_test.shape[2]).transpose()
        videoLabel_test = intList2Array(videoLabel_test)
        frameLabel_test = list2Array(frameLabel_test, 0)
        clipMarker_test = list2Array(clipMarker_test, 0)

        
        # Transition parameter
        sampleInd_train_valid = np.concatenate((sampleInd_train, sampleInd_valid))
        
        frameLabel_trans = getSample(frameLabel, sampleInd_train_valid)
                
        priors, transmat = getTransmat(frameLabel_trans, stateTypeNum_all)
            

        scalar = None
        if normalizationType == 2:
            sample_train_valid = np.concatenate((sample_train, sample_valid))
            
            scalar = preprocessing.StandardScaler().fit(sample_train_valid)

        if normalizationType == 1:
            scalar = preprocessing.StandardScaler().fit(sample_train)


        sample_train = scalar.transform(sample_train)
        sample_valid = scalar.transform(sample_valid)
        #sample_test = scalar.transform(sample_test)




        

        
        
        dataPath_train = os.path.join('Data', 'train.pkl')
        dataPath_valid = os.path.join('Data', 'valid.pkl')
        dataPath_test = os.path.join('Data', 'test.pkl')
        
        f = open(dataPath_train,'wb')
        pickle.dump( {"sample": sample_train, 
                      "videoLabel": videoLabel_train, 
                      "frameLabel": frameLabel_train,
                      "clipMarker": clipMarker_train},
                    f) 
        f.close() 
        
        
        
        
        f = open(dataPath_valid,'wb')
        pickle.dump( {"sample": sample_valid, 
                      "videoLabel": videoLabel_valid,  
                      "frameLabel": frameLabel_valid,
                      "clipMarker": clipMarker_valid},
                    f)
        f.close() 
        
        f = open(dataPath_test,'wb')
        pickle.dump( {"sample": sample_test, 
                      "videoLabel": videoLabel_test, 
                      "frameLabel": frameLabel_test,
                      "clipMarker": clipMarker_test},
                    f)
        f.close()   
        
        
        
        netPath_train = os.path.join(currentWorkingDir, "net_train.prototxt")
        netPath_valid = os.path.join(currentWorkingDir, "net_valid.prototxt")
        netPath_test = os.path.join(currentWorkingDir, "net_test.prototxt")
        solverPath = os.path.join(currentWorkingDir, "solver.prototxt")
        
        
        
        dataFileParam_train = dict(data_path = dataPath_train, batch_size = batchSize)
        dataFileParam_valid = dict(data_path = dataPath_valid, batch_size = batchSize)
        dataFileParam_test = dict(data_path = dataPath_test, batch_size = batchSize)
    
        net_train = defineNet('train', batchSize, dataFileParam_train, stateTypeNumAll, sample_train.shape[1])
        net_valid = defineNet('test', batchSize, dataFileParam_valid, stateTypeNumAll, sample_train.shape[1])
        net_test = defineNet('test', batchSize, dataFileParam_test, stateTypeNumAll, sample_train.shape[1])
        

        solver = defineSolver(train_net_path = netPath_train, 
                              test_net_path = netPath_valid, 
                              base_lr = 0.01,
                              snapshot_dest = './')
        
        with open(netPath_train, 'w') as f:
            f.write(str(net_train.to_proto()))
          
        with open(netPath_valid, 'w') as f:   
            f.write(str(net_valid.to_proto()))
                
        with open(netPath_test, 'w') as f:   
            f.write(str(net_test.to_proto()))
        
        with open(solverPath, 'w') as f:
            f.write(str(solver))
    
        

        
        
        data_valid = {'sample': sample_valid,
                      'videoLabel': videoLabel_valid,
                      'frameLabel': frameLabel_valid,
                      'clipMarker': clipMarker_valid}
        
        
        currentWeightDir = os.path.join(currentWorkingDir, 'weights')
        myMkdir(currentWeightDir)
        
        validAccuracy, \
        weightPath, \
        validLoss = train(currentWeightDir, 
                          solverPath, 
                          netPath_train, 
                          netPath_valid, 
                          batchSize, 
                          sample_train.shape[0], 
                          sample_valid.shape[0], 
                          trainingEpoches, 
                          data_valid)
    
        currentBestWeightInd = validAccuracy.argmax()
        currentBestValidAccuracy = validAccuracy[currentBestWeightInd]
        currentBestWeightPath = weightPath[currentBestWeightInd]
        
        
        


        
        
        
        predcitions = testOneNet(netPath_test, 
                                           currentBestWeightPath, 
                                           transmat, 
                                           priors, 
                                           data_test['sample'], 
                                           stateTypeNumAll, 
                                           batchSize,
                                           scalar)

        
        currentTestAccuracy = np.zeros(shape = validAccuracy.shape, dtype = np.float)
        

            
        currentSamplingResult = mSampler.getNextSamplingResult()
        workingDirInd += 1

def getTransmat(frameLabel, stateTypeNum_all):
    
    prior = np.zeros(shape=(stateTypeNum_all))
    transmat = np.zeros(shape=(stateTypeNum_all,stateTypeNum_all))
    
    for i in range(len(frameLabel)):
        
        currentFrameLabel = frameLabel[i].astype('int')
        prior[currentFrameLabel[0]] += 1
        
        for j in range(currentFrameLabel.shape[0]-1):
            transmat[currentFrameLabel[j], currentFrameLabel[j+1]] += 1
    
    return prior, transmat
    
def createClipMarkers(label):
    
    out = list()
    
    for i in range(len(label)):
        currentLabel = label[i]
        
        frameLength = currentLabel.shape[0]
        
        currentClipMakers = np.ones([frameLength, 1])
        currentClipMakers[0] = 0
        out.append(currentClipMakers)
        
    return out
        
def intList2Array (list_):      
    
    output = np.zeros([len(list_), 1])
    
    for i in range(len(list_)):
        output[i] = int(list_[i])     
        
    return output   
    
def list2Array(list_, axis_):
    
    array_ = list_[0] 
    
    
    for i in range(1, len(list_)):
        array_ = np.concatenate((array_, list_[i]), axis = axis_)
          
    return array_

    
def getSample(samples, ind):
    ind = ind.astype('int')
    output = list()
    for i in range(len(ind)):
        output.append(samples[ind[i]])
          
    return output
            
def getMaxLength(sample):
    lengths = np.zeros(len(sample))
        
    for i in range(len(sample)):
        lengths[i] = sample[i].shape[2]
            
    return max(lengths)


if __name__ == '__main__':
    main()